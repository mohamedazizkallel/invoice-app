from django.contrib.auth.decorators import login_required,user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.http import url_has_allowed_host_and_scheme
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q,Sum, Count
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from django.conf import settings
from django.utils import timezone
from django.http import JsonResponse
from .forms import ClientForm, InvoiceForm, SupplierForm, UserLoginForm, SettingsForm, ServiceForm, ClientTransactionForm, SupplierTransactionForm, SupplyForm, PurchaseForm
from .models import Client,Invoice,Settings,Service,InvoiceService,Supplier,ClientTransaction, SupplierTransaction, Supply, Purchase, PurchaseLine, InvoiceSupplyUsage
from payment.models import Retenu, PurchaseRetenu
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,logout,login as auth_login
from random import randint
from uuid import uuid4
import json
from num2words import num2words
from decimal import Decimal
from .utilities import num2words_tnd_fr

def anonymous_required(function=None, redirect_url=None):
    if not redirect_url:
        redirect_url="dashboard"

    actual_decorator = user_passes_test(
        lambda u: u.is_anonymous,
        login_url=redirect_url
    )

    if function:
        return actual_decorator(function)
    return actual_decorator

@anonymous_required
def login_view(request):  # changed name
    context = {}

    if request.method == 'GET':
        form = UserLoginForm()
        context['form'] = form
        return render(request, 'sales/login.html', context)

    if request.method == 'POST':
        form = UserLoginForm(request.POST)

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            auth_login(request, user)  # use Django's login
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid Credentials')
            return render(request, 'sales/login.html', {'form': form})
        
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    # Get recent invoices
    invoices = Invoice.objects.all().select_related('client').prefetch_related('service').order_by('-date_created')[:10]

    # Calculate statistics
    total_invoices = Invoice.objects.count()

    # Outstanding amount (CURRENT + OVERDUE invoices)
    outstanding_invoices = Invoice.objects.filter(status__in=['CURRENT', 'OVERDUE'])
    outstanding_amount = sum(invoice.calculate_total() for invoice in outstanding_invoices)

    # Paid this month
    from datetime import datetime
    current_month = datetime.now().month
    current_year = datetime.now().year
    paid_this_month_invoices = Invoice.objects.filter(
        status='PAID',
        date_created__month=current_month,
        date_created__year=current_year
    )
    paid_this_month_amount = sum(invoice.calculate_total() for invoice in paid_this_month_invoices)

    # Currency (default to TND)
    currency = 'TND'

    context = {
        'invoices': invoices,
        'total_invoices': total_invoices,
        'outstanding_amount': outstanding_amount,
        'paid_this_month_amount': paid_this_month_amount,
        'currency': currency,
    }
    return render(request,"sales/dashboard.html", context)


@login_required
def clients(request):
    clients_qs = Client.objects.all()
    
    if request.method == 'POST':
        form = ClientForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'New Client Added')
            return redirect('clients')
        else:
            messages.error(request, 'Problem processing your request')
            return redirect('clients')
    
    form = ClientForm()
    transaction_form = ClientTransactionForm()
    return render(request, 'sales/clients.html', {
        'clients': clients_qs,
        'form': form,
        'transaction_form': transaction_form,
    })

@login_required
def edit_client(request, client_id):
    """Edit an existing service"""
    client = get_object_or_404(Client, id=client_id)
    
    if request.method == 'POST':
        # Manually handle form data
        clientname = request.POST.get('clientname')
        emailAddress = request.POST.get('emailAddress')
        adress = request.POST.get('adress')
        mf = request.POST.get('mf')
        
        try:
            # Update service fields
            client.clientname = clientname
            client.emailAddress = emailAddress if emailAddress else ''
            client.adress = adress if adress else ''
            client.mf = mf if mf else ''            

            client.save()
            messages.success(request, f'Client "{client.clientname}" updated successfully!')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid data provided: {str(e)}')
    
    return redirect('clients')

@login_required
def delete_client(request, pk):
    client = get_object_or_404(Client, pk=pk)
    client.delete()
    messages.success(request, "Client removed successfully")
    return redirect('clients')


@login_required
def client_transactions(request, client_id):
    """AJAX endpoint: return client transactions and balance as JSON"""
    client = get_object_or_404(Client, id=client_id)
    transactions = client.transactions.all().order_by('date_created')

    running_balance = Decimal('0')
    transaction_list = []
    for txn in transactions:
        if txn.transaction_type == 'DEBIT':
            running_balance += txn.amount
        else:
            running_balance -= txn.amount

        transaction_list.append({
            'id': txn.id,
            'date': txn.date_created.strftime('%Y-%m-%d %H:%M') if txn.date_created else '',
            'type': txn.transaction_type,
            'source': txn.get_source_display(),
            'description': txn.description or '',
            'amount': float(txn.amount),
            'running_balance': float(running_balance),
            'invoice_id': txn.invoice_id,
            'invoice_ref': txn.invoice.uniqueId if txn.invoice else None,
        })

    return JsonResponse({
        'success': True,
        'client_name': client.clientname,
        'balance': float(client.get_balance()),
        'transactions': transaction_list,
    })


@login_required
def client_add_transaction(request, client_id):
    """Add a manual credit/debit entry for a client"""
    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        form = ClientTransactionForm(request.POST)
        if form.is_valid():
            txn = form.save(commit=False)
            txn.client = client
            txn.source = 'MANUAL'
            txn.save()
            messages.success(request, f'Transaction ajoutée pour "{client.clientname}"')
        else:
            messages.error(request, 'Données de transaction invalides')

    return redirect('clients')

@login_required
def suppliers(request):
    suppliers_qs = Supplier.objects.all()
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'New Supplier Added')
            return redirect('suppliers')
        else:
            messages.error(request, 'Problem processing your request')
            return redirect('suppliers')
    
    form = SupplierForm()
    transaction_form = SupplierTransactionForm()
    return render(request, 'sales/supplier.html', {'Suppliers': suppliers_qs, 'form': form, 'transaction_form': transaction_form})

@login_required
def edit_supplier(request, client_id):
    """Edit an existing service"""
    supplier = get_object_or_404(Supplier, id=client_id)
    
    if request.method == 'POST':
        # Manually handle form data
        name = request.POST.get('name')
        emailAddress = request.POST.get('emailAddress')
        adress = request.POST.get('adress')
        mf = request.POST.get('mf')
        
        try:
            # Update service fields
            supplier.name = name
            supplier.emailAddress = emailAddress if emailAddress else ''
            supplier.adress = adress if adress else ''
            supplier.mf = mf if mf else ''            

            supplier.save()
            messages.success(request, f'Supplier "{supplier.name}" updated successfully!')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid data provided: {str(e)}')
    
    return redirect('suppliers')

@login_required
def delete_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    supplier.delete()
    messages.success(request, "Supplier removed successfully")
    return redirect('suppliers')


@login_required
def supplier_transactions(request, supplier_id):
    """AJAX endpoint: return supplier transactions and balance as JSON"""
    supplier = get_object_or_404(Supplier, id=supplier_id)
    transactions = supplier.supplier_transactions.all().order_by('date_created')

    running_balance = Decimal('0')
    transaction_list = []
    for txn in transactions:
        if txn.transaction_type == 'CREDIT':
            running_balance += txn.amount
        else:
            running_balance -= txn.amount

        transaction_list.append({
            'id': txn.id,
            'date': txn.date_created.strftime('%Y-%m-%d %H:%M') if txn.date_created else '',
            'type': txn.transaction_type,
            'source': txn.get_source_display(),
            'description': txn.description or '',
            'amount': float(txn.amount),
            'running_balance': float(running_balance),
            'purchase_id': txn.purchase_id,
            'purchase_ref': txn.purchase.uniqueId if txn.purchase else None,
        })

    return JsonResponse({
        'success': True,
        'supplier_name': supplier.name,
        'balance': float(supplier.get_balance()),
        'transactions': transaction_list,
    })


@login_required
def supplier_add_transaction(request, supplier_id):
    """Add a manual credit/debit entry for a supplier"""
    supplier = get_object_or_404(Supplier, id=supplier_id)

    if request.method == 'POST':
        form = SupplierTransactionForm(request.POST)
        if form.is_valid():
            txn = form.save(commit=False)
            txn.supplier = supplier
            txn.source = 'MANUAL'
            txn.save()
            messages.success(request, f'Transaction ajoutée pour "{supplier.name}"')
        else:
            messages.error(request, 'Données de transaction invalides')

    return redirect('suppliers')


# ============ SUPPLIES CRUD ============

@login_required
def supplies_list(request):
    """Display all supplies"""
    supplies = Supply.objects.all().select_related('preferred_supplier')

    search_query = request.GET.get('search', '')
    if search_query:
        supplies = supplies.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__icontains=search_query)
        )

    form = SupplyForm()
    suppliers_qs = Supplier.objects.all().order_by('name')
    return render(request, 'sales/supplies.html', {
        'supplies': supplies,
        'form': form,
        'suppliers': suppliers_qs,
    })


@login_required
def supply_create(request):
    """Create a new supply"""
    if request.method == 'POST':
        form = SupplyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fourniture ajoutée avec succès')
        else:
            messages.error(request, 'Erreur lors de l\'ajout de la fourniture')
    return redirect('supplies_list')


@login_required
def supply_edit(request, supply_id):
    """Edit an existing supply"""
    supply = get_object_or_404(Supply, id=supply_id)

    if request.method == 'POST':
        name = request.POST.get('name')
        category = request.POST.get('category')
        unit = request.POST.get('unit')
        unit_price = request.POST.get('unit_price')
        stock_quantity = request.POST.get('stock_quantity')
        min_stock = request.POST.get('min_stock')
        preferred_supplier_id = request.POST.get('preferred_supplier')
        description = request.POST.get('description')

        try:
            supply.name = name
            supply.category = category if category else 'raw_material'
            supply.unit = unit if unit else 'pièce'
            supply.unit_price = Decimal(unit_price) if unit_price else Decimal('0')
            supply.stock_quantity = Decimal(stock_quantity) if stock_quantity else Decimal('0')
            supply.min_stock = Decimal(min_stock) if min_stock else Decimal('0')
            supply.description = description if description else ''

            if preferred_supplier_id:
                supply.preferred_supplier = Supplier.objects.get(id=preferred_supplier_id)
            else:
                supply.preferred_supplier = None

            supply.save()
            messages.success(request, f'Fourniture "{supply.name}" modifiée avec succès')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Données invalides : {str(e)}')

    return redirect('supplies_list')


@login_required
def supply_delete(request, supply_id):
    """Delete a supply"""
    supply = get_object_or_404(Supply, id=supply_id)
    if request.method == 'POST':
        supply_name = supply.name
        supply.delete()
        messages.success(request, f'Fourniture "{supply_name}" supprimée avec succès')
    return redirect('supplies_list')


# ============ PURCHASES CRUD ============

@login_required
def purchases_list(request):
    """Display all purchases"""
    purchases = Purchase.objects.all().select_related('supplier')

    search_query = request.GET.get('search', '')
    if search_query:
        purchases = purchases.filter(
            Q(uniqueId__icontains=search_query) |
            Q(supplier__name__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    status_filter = request.GET.get('status', '')
    if status_filter:
        purchases = purchases.filter(status=status_filter)

    suppliers_qs = Supplier.objects.all().order_by('name')
    supplies = Supply.objects.all().order_by('name')

    context = {
        'purchases': purchases,
        'suppliers': suppliers_qs,
        'supplies': supplies,
        'form': PurchaseForm(),
    }
    return render(request, 'sales/purchases.html', context)


@login_required
def purchase_create(request):
    """Create a new purchase with line items"""
    if request.method != 'POST':
        return redirect('purchases_list')

    try:
        with transaction.atomic():
            supplier_id = request.POST.get('supplier')
            if not supplier_id:
                messages.error(request, 'Le fournisseur est requis.')
                return redirect('purchases_list')

            try:
                supplier = Supplier.objects.get(id=supplier_id)
            except Supplier.DoesNotExist:
                messages.error(request, 'Fournisseur introuvable.')
                return redirect('purchases_list')

            notes = request.POST.get('notes', '')
            tva = request.POST.get('tva', '19.00').strip()
            discount = request.POST.get('discount', '0.00').strip()
            timbre = request.POST.get('timbre_fiscal', '1.000').strip()

            purchase = Purchase.objects.create(
                supplier=supplier,
                notes=notes,
                tva=Decimal(tva) if tva else Decimal('19.00'),
                discount=Decimal(discount) if discount else Decimal('0.00'),
                timbre_fiscal=Decimal(timbre) if timbre else Decimal('1.000'),
            )

            supply_ids = request.POST.getlist('supply_id[]')
            quantities = request.POST.getlist('quantity[]')
            unit_prices = request.POST.getlist('line_unit_price[]')

            if not supply_ids:
                messages.error(request, 'Vous devez ajouter au moins une ligne.')
                purchase.delete()
                return redirect('purchases_list')

            for i, supply_id in enumerate(supply_ids):
                if not supply_id:
                    continue
                try:
                    supply = Supply.objects.get(id=supply_id)
                    qty = Decimal(quantities[i]) if i < len(quantities) and quantities[i] else Decimal('1')
                    price = Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else supply.unit_price
                    PurchaseLine.objects.create(
                        purchase=purchase,
                        supply=supply,
                        quantity=qty,
                        unit_price=price,
                    )
                except Supply.DoesNotExist:
                    pass

            messages.success(request, f'Achat #{purchase.uniqueId} créé avec succès.')
            return redirect('purchase_detail', purchase.id)

    except Exception as e:
        messages.error(request, f'Erreur lors de la création de l\'achat : {str(e)}')

    return redirect('purchases_list')


@login_required
def purchase_detail(request, purchase_id):
    """View purchase details"""
    purchase = get_object_or_404(Purchase, id=purchase_id)
    purchase_lines = purchase.purchase_lines.select_related('supply').all()

    subtotal = purchase.calculate_subtotal()
    discount_amount = purchase.calculate_discount_amount()
    subtotal_after_discount = purchase.calculate_subtotal_after_discount()
    tva_amount = purchase.calculate_tva_amount()
    total = purchase.calculate_total()
    total_retenue = purchase.get_total_retenue()
    net_amount = purchase.get_net_amount()

    retenu_types = Retenu.objects.filter(is_active=True).order_by('category', 'rate')
    purchase_retenues = purchase.purchase_retenues.select_related('retenu_type').all()

    suppliers_qs = Supplier.objects.all().order_by('name')
    supplies = Supply.objects.all().order_by('name')

    context = {
        'purchase': purchase,
        'purchase_lines': purchase_lines,
        'subtotal': subtotal,
        'discount_amount': discount_amount,
        'subtotal_after_discount': subtotal_after_discount,
        'tva_amount': tva_amount,
        'total': total,
        'total_retenue': total_retenue,
        'net_amount': net_amount,
        'retenu_types': retenu_types,
        'purchase_retenues': purchase_retenues,
        'suppliers': suppliers_qs,
        'supplies': supplies,
    }
    return render(request, 'sales/purchase_detail.html', context)


@login_required
def purchase_edit(request, purchase_id):
    """Edit an existing purchase"""
    purchase = get_object_or_404(Purchase, id=purchase_id)

    if request.method != 'POST':
        return redirect('purchase_detail', purchase.id)

    try:
        with transaction.atomic():
            supplier_id = request.POST.get('supplier')
            if supplier_id:
                try:
                    purchase.supplier = Supplier.objects.get(id=supplier_id)
                except Supplier.DoesNotExist:
                    raise ValueError('Fournisseur introuvable.')

            purchase.notes = request.POST.get('notes', '')
            if request.POST.get('tva'):
                purchase.tva = Decimal(request.POST['tva'])
            if request.POST.get('discount'):
                purchase.discount = Decimal(request.POST['discount'])
            if request.POST.get('timbre_fiscal'):
                purchase.timbre_fiscal = Decimal(request.POST['timbre_fiscal'])

            # Rebuild lines
            purchase.purchase_lines.all().delete()

            supply_ids = request.POST.getlist('supply_id[]')
            quantities = request.POST.getlist('quantity[]')
            unit_prices = request.POST.getlist('line_unit_price[]')

            if not supply_ids:
                raise ValueError('Vous devez ajouter au moins une ligne.')

            for i, supply_id in enumerate(supply_ids):
                if not supply_id:
                    continue
                try:
                    supply = Supply.objects.get(id=supply_id)
                    qty = Decimal(quantities[i]) if i < len(quantities) and quantities[i] else Decimal('1')
                    price = Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else supply.unit_price
                    PurchaseLine.objects.create(
                        purchase=purchase,
                        supply=supply,
                        quantity=qty,
                        unit_price=price,
                    )
                except Supply.DoesNotExist:
                    pass

            purchase.save()
            messages.success(request, f'Achat #{purchase.uniqueId} modifié avec succès.')

    except (ValueError, TypeError) as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Erreur : {str(e)}')

    return redirect('purchase_detail', purchase.id)


@login_required
def purchase_delete(request, purchase_id):
    """Delete a purchase and reverse stock/ledger if confirmed"""
    purchase = get_object_or_404(Purchase, id=purchase_id)

    if request.method == 'POST':
        purchase_ref = purchase.uniqueId

        # If purchase was received, reverse stock
        if purchase.status in ('RECEIVED', 'PAID'):
            for line in purchase.purchase_lines.all():
                line.supply.stock_quantity -= line.quantity
                line.supply.save()

            # Reverse ledger entries
            if purchase.supplier:
                purchase_total = purchase.calculate_total()
                SupplierTransaction.objects.create(
                    supplier=purchase.supplier,
                    purchase=None,
                    transaction_type='DEBIT',
                    source='PURCHASE_DELETED',
                    amount=purchase_total,
                    description=f'Annulation achat #{purchase_ref}'
                )

        purchase.delete()
        messages.success(request, f'Achat #{purchase_ref} supprimé avec succès.')

    return redirect('purchases_list')


@login_required
def purchase_confirm(request, purchase_id):
    """Confirm/receive a purchase: increment stock and create supplier CREDIT"""
    purchase = get_object_or_404(Purchase, id=purchase_id)

    if request.method == 'POST':
        if purchase.status not in ('DRAFT', 'CONFIRMED'):
            messages.warning(request, 'Cet achat a déjà été reçu.')
            return redirect('purchase_detail', purchase.id)

        with transaction.atomic():
            # Increment stock for each line
            for line in purchase.purchase_lines.all():
                line.supply.stock_quantity += line.quantity
                line.supply.save()

            # Create supplier CREDIT transaction (we owe them)
            purchase_total = purchase.calculate_total()
            SupplierTransaction.objects.create(
                supplier=purchase.supplier,
                purchase=purchase,
                transaction_type='CREDIT',
                source='PURCHASE_CONFIRMED',
                amount=purchase_total,
                description=f'Achat #{purchase.uniqueId} reçu'
            )

            purchase.status = 'RECEIVED'
            purchase.save()

        messages.success(request, f'Achat #{purchase.uniqueId} confirmé et stock mis à jour.')

    return redirect('purchase_detail', purchase.id)


@login_required
def process_purchase_payment(request, purchase_id):
    """Process payment for a purchase: create supplier DEBIT"""
    purchase = get_object_or_404(Purchase, id=purchase_id)

    if request.method == 'POST':
        if purchase.status == 'PAID':
            messages.warning(request, 'Cet achat est déjà payé.')
            return redirect('purchase_detail', purchase.id)

        with transaction.atomic():
            net_amount = purchase.get_net_amount()
            SupplierTransaction.objects.create(
                supplier=purchase.supplier,
                purchase=purchase,
                transaction_type='DEBIT',
                source='PURCHASE_PAID',
                amount=net_amount,
                description=f'Paiement achat #{purchase.uniqueId}'
            )

            purchase.status = 'PAID'
            purchase.save()

        messages.success(request, f'Paiement de l\'achat #{purchase.uniqueId} enregistré.')

    return redirect('purchase_detail', purchase.id)


@login_required
def purchase_retenu_create(request, purchase_id):
    """Add retenu to a purchase"""
    purchase = get_object_or_404(Purchase, id=purchase_id)

    if request.method == 'POST':
        retenu_type_id = request.POST.get('retenu_type')
        base_amount = request.POST.get('base_amount')

        if retenu_type_id and base_amount:
            try:
                retenu_type = Retenu.objects.get(id=retenu_type_id)
                base = Decimal(base_amount)
                calculated = (base * retenu_type.rate) / Decimal('100')

                PurchaseRetenu.objects.create(
                    purchase=purchase,
                    retenu_type=retenu_type,
                    base_amount=base,
                    retenu_rate=retenu_type.rate,
                    retenu_amount=calculated,
                )
                messages.success(request, 'Retenue ajoutée avec succès.')
            except (Retenu.DoesNotExist, ValueError) as e:
                messages.error(request, f'Erreur : {str(e)}')
        else:
            messages.error(request, 'Données manquantes.')

    return redirect('purchase_detail', purchase.id)


@login_required
def purchase_retenu_delete(request, retenu_id):
    """Delete a retenu from a purchase"""
    retenu = get_object_or_404(PurchaseRetenu, id=retenu_id)
    purchase_id = retenu.purchase_id

    if request.method == 'POST':
        retenu.delete()
        messages.success(request, 'Retenue supprimée.')

    return redirect('purchase_detail', purchase_id)


@login_required
def invoices_list(request):
    """Display all invoices with filtering and search"""
    invoices = Invoice.objects.all().select_related('client').prefetch_related('service')
    
    # Search filter
    search_query = request.GET.get('search', '')
    if search_query:
        invoices = invoices.filter(
            Q(title__icontains=search_query) | 
            Q(client__clientname__icontains=search_query) |
            Q(notes__icontains=search_query) |
            Q(uniqueId__icontains=search_query)
        )
    
    # Status filter
    status = request.GET.get('status', '')
    if status:
        invoices = invoices.filter(status=status)
    
    # Client filter
    client_id = request.GET.get('client', '')
    if client_id:
        invoices = invoices.filter(client_id=client_id)
    
    # Date filter
    date_from = request.GET.get('date_from', '')
    if date_from:
        invoices = invoices.filter(date_created__gte=date_from)
    
    # Sorting
    sort_by = request.GET.get('sort', '-date_created')
    invoices = invoices.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(invoices, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all clients and services for dropdowns
    clients = Client.objects.all().order_by('clientname')
    services = Service.objects.all().order_by('title')
    
    # Get Settings for form defaults (create if doesn't exist)
    settings = Settings.objects.first()
    if not settings:
        settings = Settings.objects.create(
            clientname="My Company",
            tva=Decimal('19.00'),
            dt=Decimal('1.000')
        )

    retenu_types = Retenu.objects.filter(is_active=True).order_by('category', 'rate')

    # Calculate statistics
    total_invoices = invoices.count()
    current_invoices = invoices.filter(status='CURRENT').count()
    overdue_invoices = invoices.filter(status='OVERDUE').count()
    paid_invoices = invoices.filter(status='PAID').count()

    context = {
        'invoices': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'clients': clients,
        'services': services,
        'form': InvoiceForm(),
        'total_invoices': total_invoices,
        'current_invoices': current_invoices,
        'overdue_invoices': overdue_invoices,
        'paid_invoices': paid_invoices,
        'settings': settings,
        'retenu_types': retenu_types,
        'current_year': timezone.now().year,
    }
    
    return render(request, 'sales/invoice_service.html', context)

@login_required
def invoice_create(request):
    """Create a new invoice with inventory management"""
    if request.method != 'POST':
        return redirect('invoices_list')

    try:
        with transaction.atomic():
            # Required fields
            client_id = request.POST.get('client')
            title = ''  # Title is now optional/empty

            if not client_id:
                messages.error(request, 'Client is required.')
                return redirect('invoices_list')

            try:
                client = Client.objects.get(id=client_id)
            except Client.DoesNotExist:
                messages.error(request, 'Selected client does not exist.')
                return redirect('invoices_list')

            # Basic values
            status = request.POST.get('status', 'CURRENT')
            notes = request.POST.get('notes', '')
            # Get Settings for defaults
            settings = Settings.objects.first()
            
            # TVA and Timbre Fiscal - use form values or fall back to Settings
            tva_input = request.POST.get('tva', '').strip()
            if tva_input:
                tva = float(tva_input)
            elif settings and settings.tva:
                tva = float(settings.tva)
            else:
                tva = 19.00
            
            timbre_input = request.POST.get('timbre_fiscal', '').strip()
            if timbre_input:
                timbre_fiscal = float(timbre_input)
            elif settings and settings.dt:
                timbre_fiscal = float(settings.dt)
            else:
                timbre_fiscal = 1.000
            
            discount = float(request.POST.get('discount', 0.00))

            # Get services from arrays
            service_ids = request.POST.getlist('service_id[]')

            if not service_ids:
                messages.error(request, 'You must add at least one service.')
                return redirect('invoices_list')

            # Create invoice
            invoice = Invoice.objects.create(
                title=title,
                client=client,
                status=status,
                notes=notes,
                tva=tva,
                timbre_fiscal=timbre_fiscal,
                discount=discount
            )

            # Add services (if you have a Service model and InvoiceService model)
            fodec_flags = request.POST.getlist('has_fodec[]')
            unit_prices = request.POST.getlist('unit_price[]')
            for i, service_id in enumerate(service_ids):
                if not service_id:
                    continue

                try:
                    service = Service.objects.get(id=service_id)
                    has_fodec = fodec_flags[i] == '1' if i < len(fodec_flags) else False
                    # Use submitted price if provided, otherwise fall back to service price
                    if i < len(unit_prices) and unit_prices[i]:
                        price = Decimal(str(unit_prices[i]))
                    else:
                        price = service.price
                    InvoiceService.objects.create(
                        invoice=invoice,
                        service=service,
                        unit_price=price,
                        has_fodec=has_fodec
                    )
                except Service.DoesNotExist:
                    pass  # Skip if service doesn't exist

            # Auto-DEBIT ledger entry
            invoice_total = invoice.calculate_total()
            ClientTransaction.objects.create(
                client=client,
                invoice=invoice,
                transaction_type='DEBIT',
                source='INVOICE_CREATED',
                amount=invoice_total,
                description=f'Facture {invoice.uniqueId} - {invoice.title}'
            )

            messages.success(request, f'Invoice "{invoice.title}" created successfully.')
            return redirect('invoice_detail', invoice.id)

    except Service.DoesNotExist:
        messages.error(request, 'One of the selected services does not exist.')
    except ValueError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Error creating invoice: {str(e)}')

    return redirect('invoices_list')

@login_required
def invoice_edit(request, invoice_id):
    """Edit an existing invoice and adjust inventory properly"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    next_page = request.POST.get('next', 'detail')

    if request.method != 'POST':
        return redirect('invoice_detail', invoice.id)

    try:
        with transaction.atomic():
            # Basic fields
            title = ''  # Title is now optional/empty
            status = request.POST.get('status')
            notes = request.POST.get('notes', '')

            invoice.title = title
            invoice.status = status
            invoice.notes = notes

            # Numeric fields
            if request.POST.get('tva'):
                invoice.tva = float(request.POST['tva'])

            if request.POST.get('timbre_fiscal'):
                invoice.timbre_fiscal = float(request.POST['timbre_fiscal'])

            if request.POST.get('discount'):
                invoice.discount = float(request.POST['discount'])

            # Client
            client_id = request.POST.get('client')
            if client_id:
                try:
                    invoice.client = Client.objects.get(id=client_id)
                except Client.DoesNotExist:
                    raise ValueError('Selected client does not exist.')


            # Get services from arrays
            service_ids = request.POST.getlist('service_id[]')

            if not service_ids:
                raise ValueError('You must add at least one service.')
            # Delete previous invoice services and services
            invoice.invoice_services.all().delete()
            # If you have services: invoice.invoice_services.all().delete()

            # Add new services
            fodec_flags = request.POST.getlist('has_fodec[]')
            unit_prices = request.POST.getlist('unit_price[]')
            for i, service_id in enumerate(service_ids):
                if not service_id:
                    continue

                try:
                    service = Service.objects.get(id=service_id)
                    has_fodec = fodec_flags[i] == '1' if i < len(fodec_flags) else False
                    if i < len(unit_prices) and unit_prices[i]:
                        price = Decimal(str(unit_prices[i]))
                    else:
                        price = service.price
                    InvoiceService.objects.create(
                        invoice=invoice,
                        service=service,
                        unit_price=price,
                        has_fodec=has_fodec
                    )
                except Service.DoesNotExist:
                    pass

            invoice.save()

            # Update ledger: adjust DEBIT amount to new total
            new_total = invoice.calculate_total()
            if invoice.client:
                debit_entry = ClientTransaction.objects.filter(
                    invoice=invoice,
                    source='INVOICE_CREATED',
                    transaction_type='DEBIT'
                ).first()

                if debit_entry:
                    debit_entry.amount = new_total
                    debit_entry.description = f'Facture {invoice.uniqueId} - {invoice.title}'
                    debit_entry.save()
                else:
                    # Invoice existed before ledger system
                    ClientTransaction.objects.create(
                        client=invoice.client,
                        invoice=invoice,
                        transaction_type='DEBIT',
                        source='INVOICE_CREATED',
                        amount=new_total,
                        description=f'Facture {invoice.uniqueId} - {invoice.title}'
                    )

                # Auto-CREDIT when status changes to PAID
                if status == 'PAID':
                    existing_credit = ClientTransaction.objects.filter(
                        invoice=invoice,
                        source='INVOICE_PAID',
                        transaction_type='CREDIT'
                    ).exists()
                    if not existing_credit:
                        ClientTransaction.objects.create(
                            client=invoice.client,
                            invoice=invoice,
                            transaction_type='CREDIT',
                            source='INVOICE_PAID',
                            amount=new_total,
                            description=f'Paiement facture {invoice.uniqueId} - {invoice.title}'
                        )

            messages.success(request, f'Invoice "{invoice.title}" updated successfully.')

    except (ValueError, TypeError) as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Error updating invoice: {str(e)}')

    if next_page == 'detail':
        return redirect('invoice_detail', invoice.id)

    return redirect('invoices_list')

@login_required
def invoice_detail(request, invoice_id=None, slug=None):
    """View invoice details with inventory-tracked services"""
    if slug:
        invoice = get_object_or_404(Invoice, slug=slug)
    else:
        invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Get invoice services with their quantities
    invoice_services = invoice.invoice_services.select_related('service').all()
    
    # Calculate amounts
    subtotal = invoice.calculate_service_subtotal()
    discount_amount = invoice.calculate_discount_amount()
    subtotal_after_discount = invoice.calculate_subtotal_after_discount()
    total_fodec = invoice.calculate_total_fodec()
    tva_amount = invoice.calculate_tva_amount()
    total = invoice.calculate_total()
    total_in_words = num2words_tnd_fr(Decimal(total))
    # Prepare services with line totals
    services_with_totals = []
    invoice_currency = 'TND'
    

    for invoice_service in invoice_services:
        service = invoice_service.service
        if not invoice_currency or invoice_currency == 'TND':
            invoice_currency = service.currency or 'TND'

        services_with_totals.append({
            'service': service,
            'unit_price': invoice_service.unit_price,
            'line_total': invoice_service.get_line_ht(),
            'has_fodec': invoice_service.has_fodec,
            'fodec_amount': invoice_service.get_fodec_amount(),
        })

    
    # Get all clients and services for edit modal
    clients = Client.objects.all().order_by('clientname')
    all_services = Service.objects.all().order_by('title')
    # Get settings
    try:
        p_settings = Settings.objects.first()
    except Exception as e:
        p_settings = None
    
    context = {
        'invoice': invoice,
        'invoice_services': invoice_services,
        'services_with_totals': services_with_totals,
        'subtotal': subtotal,
        'discount_amount': discount_amount,
        'subtotal_after_discount': subtotal_after_discount,
        'total_fodec': total_fodec,
        'tva_amount': tva_amount,
        'total': total,
        'total_in_words': total_in_words,
        'invoiceCurrency': invoice_currency,
        'clients': clients,
        'all_services': all_services,
        'p_settings': p_settings,
    }
    
    return render(request, 'sales/invoice_detail_service.html', context)

@login_required
def invoice_delete(request, invoice_id):
    """Delete an invoice and restore inventory"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.method == 'POST':
        invoice_title = invoice.title

        # Create reversal ledger entries before deletion
        if invoice.client:
            invoice_total = invoice.calculate_total()

            had_debit = ClientTransaction.objects.filter(
                invoice=invoice, source='INVOICE_CREATED', transaction_type='DEBIT'
            ).exists()
            if had_debit:
                ClientTransaction.objects.create(
                    client=invoice.client,
                    invoice=None,
                    transaction_type='CREDIT',
                    source='INVOICE_DELETED',
                    amount=invoice_total,
                    description=f'Annulation facture {invoice.uniqueId} - {invoice_title}'
                )

            had_credit = ClientTransaction.objects.filter(
                invoice=invoice, source='INVOICE_PAID', transaction_type='CREDIT'
            ).exists()
            if had_credit:
                ClientTransaction.objects.create(
                    client=invoice.client,
                    invoice=None,
                    transaction_type='DEBIT',
                    source='INVOICE_DELETED',
                    amount=invoice_total,
                    description=f'Annulation paiement facture {invoice.uniqueId} - {invoice_title}'
                )

        invoice.delete()
        messages.success(request, f'Invoice "{invoice_title}" deleted and inventory restored!')

    return redirect('invoices_list')

@login_required
def export_invoices(request):
    """Export all invoices to Excel"""
    invoices = Invoice.objects.all().select_related('client', 'service').order_by('-date_created')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Define headers
    headers = ['Invoice ID', 'Unique ID', 'Title', 'Client', 'Service', 'Status', 'Date Created', 'Last Updated', 'Notes']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add invoice data
    for invoice in invoices:
        ws.append([
            invoice.id,
            invoice.uniqueId if invoice.uniqueId else '',
            invoice.title if invoice.title else '',
            invoice.client.clientname if invoice.client else 'No Client',
            invoice.service.title if invoice.service else 'No Service',
            invoice.status,
            invoice.date_created.strftime('%Y-%m-%d %H:%M') if invoice.date_created else '',
            invoice.last_updated.strftime('%Y-%m-%d %H:%M') if invoice.last_updated else '',
            invoice.notes if invoice.notes else '',
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 40
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=invoices_export.xlsx'
    
    wb.save(response)
    return response

@login_required
def download_invoice_template(request):
    """Download an Excel template for importing invoices"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices Template"
    
    # Define headers
    headers = ['Title', 'Client Name', 'Service Title', 'Status', 'Notes']
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add sample data
    ws.append([
        'Sample Invoice 1',
        'Sample Client',
        'Sample Service',
        'CURRENT',
        'This is a sample invoice note'
    ])
    ws.append([
        'Sample Invoice 2',
        'Another Client',
        'Another Service',
        'PAID',
        'Another sample note'
    ])
    
    # Add instructions
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ['Invoice Import Template - Instructions'],
        [''],
        ['Required Columns:'],
        ['1. Title - Invoice title (required)'],
        ['2. Client Name - Exact client name from your system (required)'],
        ['3. Service Title - Exact service title from your system'],
        ['4. Status - Invoice status: CURRENT, PAID, or OVERDUE'],
        ['5. Notes - Additional notes or comments'],
        [''],
        ['Important Notes:'],
        ['- Do not modify the header row'],
        ['- Title and Client Name are required'],
        ['- Client Name must match exactly with existing clients'],
        ['- service Title must match exactly with existing Services'],
        ['- Status values are case-sensitive (use UPPERCASE)'],
        ['- Default status is CURRENT if not specified'],
        ['- Unique ID and slug will be auto-generated'],
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    ws_instructions.column_dimensions['A'].width = 60
    ws_instructions['A1'].font = Font(bold=True, size=14)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=invoice_import_template.xlsx'
    
    wb.save(response)
    return response

@login_required
def import_invoices(request):
    """Import invoices from Excel file"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        update_existing = request.POST.get('update_existing') == 'on'
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload.')
            return redirect('invoices_list')
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file (.xlsx or .xls).')
            return redirect('invoices_list')
        
        # Validate file size (5MB limit)
        if excel_file.size > 5 * 1024 * 1024:
            messages.error(request, 'File size exceeds 5MB limit.')
            return redirect('invoices_list')
        
        try:
            # Load workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            
            # Validate required columns
            required_columns = ['Title', 'Client Name']
            for col in required_columns:
                if col not in headers:
                    messages.error(request, f'Missing required column: {col}')
                    return redirect('invoices_list')
            
            # Get column indices
            title_idx = headers.index('Title')
            client_name_idx = headers.index('Client Name')
            service_title_idx = headers.index('Service Title') if 'Service Title' in headers else None
            status_idx = headers.index('Status') if 'Status' in headers else None
            notes_idx = headers.index('Notes') if 'Notes' in headers else None
            
            # Process rows
            created_count = 0
            updated_count = 0
            error_count = 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Get values
                    title = row[title_idx]
                    client_name = row[client_name_idx]
                    
                    # Skip empty rows
                    if not title or not client_name:
                        continue
                    
                    # Get or validate client
                    try:
                        client = Client.objects.get(clientname=client_name)
                    except Client.DoesNotExist:
                        error_count += 1
                        print(f"Row {row_num}: Client '{client_name}' not found")
                        continue
                    
                    # Get service if specified
                    service = None
                    if service_title_idx is not None and row[service_title_idx]:
                        try:
                            service = Service.objects.get(title=row[service_title_idx])
                        except service.DoesNotExist:
                            print(f"Row {row_num}: Service '{row[service_title_idx]}' not found, skipping service")
                    
                    status = row[status_idx] if status_idx is not None and row[status_idx] else 'CURRENT'
                    notes = row[notes_idx] if notes_idx is not None and row[notes_idx] else ''
                    
                    # Validate status
                    if status not in ['CURRENT', 'OVERDUE', 'PAID']:
                        status = 'CURRENT'
                    
                    # Create or update invoice
                    if update_existing:
                        invoice, created = Invoice.objects.get_or_create(
                            title=title,
                            defaults={
                                'client': client,
                                'service': service,
                                'status': status,
                                'notes': notes,
                            }
                        )
                        
                        if not created:
                            invoice.client = client
                            invoice.service = service
                            invoice.status = status
                            invoice.notes = notes
                            invoice.save()
                            updated_count += 1
                        else:
                            created_count += 1
                    else:
                        invoice = Invoice.objects.create(
                            title=title,
                            client=client,
                            service=service,
                            status=status,
                            notes=notes,
                        )
                        created_count += 1
                        
                except Exception as e:
                    error_count += 1
                    print(f"Error processing row {row_num}: {str(e)}")
                    continue
            
            # Success message
            if created_count > 0 or updated_count > 0:
                msg_parts = []
                if created_count > 0:
                    msg_parts.append(f'{created_count} invoice(s) created')
                if updated_count > 0:
                    msg_parts.append(f'{updated_count} invoice(s) updated')
                
                success_msg = ' and '.join(msg_parts) + ' successfully!'
                messages.success(request, success_msg)
                
                if error_count > 0:
                    messages.warning(request, f'{error_count} row(s) had errors and were skipped.')
            else:
                if error_count > 0:
                    messages.error(request, f'Import failed. {error_count} row(s) had errors.')
                else:
                    messages.warning(request, 'No invoices were imported. Please check your file.')
                    
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('invoices_list')
    
    return redirect('invoices_list')
    
@login_required
def service_view(request):
    """Display all services with filtering and search"""
    services = Service.objects.all()
    
    # Search filter
    search_query = request.GET.get('search', '')
    if search_query:
        services = services.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    
    # Category filter (if you have categories)
    category = request.GET.get('category', '')
    if category:
        services = services.filter(category_id=category)
    
    # Sorting
    sort_by = request.GET.get('sort', 'title')
    services = services.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(services, 20)  # 20 services per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get categories for filter dropdown (if applicable)
    # categories = Category.objects.all()
    
    context = {
        'services': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'form': ServiceForm(),
        # 'categories': categories,
    }
    
    return render(request, 'sales/services.html', context)

@login_required
def add_service(request):
    """Add a single service"""
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            service = form.save(commit=False)
            # Add any additional fields if needed
            # service.created_by = request.user
            service.save()
            messages.success(request, f'Service "{service.title}" added successfully!')
            return redirect('services_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            # Return to the same page with errors
            services = Service.objects.all().order_by('title')
            context = {
                'services': services,
                'form': form,
            }
            return render(request, 'sales/services.html', context)
    
    return redirect('services_list')

@login_required
def edit_service(request, service_id):
    """Edit an existing service"""
    service = get_object_or_404(Service, id=service_id)
    
    if request.method == 'POST':
        # Manually handle form data
        title = request.POST.get('title')
        currency = request.POST.get('currency')
        description = request.POST.get('description')
        price = request.POST.get('price')
        
        try:
            # Update service fields
            service.title = title
            service.currency = currency if currency else 'TND'
            service.description = description if description else ''
            service.price = float(price) if price else 0.0
            service.apply_fodec = request.POST.get('apply_fodec') == 'on'

            service.save()
            messages.success(request, f'Service "{service.title}" updated successfully!')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid data provided: {str(e)}')
    
    return redirect('services_list')

@login_required
def delete_service(request, service_id):
    """Delete a service"""
    service = get_object_or_404(Service, id=service_id)
    
    if request.method == 'POST':
        service_title = service.title
        service.delete()
        messages.success(request, f'Service "{service_title}" deleted successfully!')
    
    return redirect('services_list')


@login_required
def settings_view(request):
    """View and edit company settings"""
    settings = Settings.objects.first()
    
    if request.method == 'POST':
        if settings:
            form = SettingsForm(request.POST, request.FILES, instance=settings)
        else:
            form = SettingsForm(request.POST, request.FILES)
        
        if form.is_valid():
            form.save()
            messages.success(request, 'Settings updated successfully!')
            return redirect('settings_view')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        if settings:
            form = SettingsForm(instance=settings)
        else:
            form = SettingsForm()
    
    context = {
        'form': form,
        'settings': settings,
    }
    
    return render(request, 'sales/settings.html', context)
